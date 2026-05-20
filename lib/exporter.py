"""Exporters for the MX configuration slices.

Emits two formats:

* ``JSONConfigExporter`` -- always runs. Produces a deterministic JSON
  payload with ``source`` metadata, ``interfaces``, and ``routing`` sections.
  The JSON shape intentionally mirrors the srx-manager shape so downstream
  tooling stays portable.
* ``ExcelConfigExporter`` -- optional. Writes a multi-sheet ``.xlsx`` with
  sheets for base interfaces, unit interfaces, routing instances, and static
  routes.

Cross-slice joins are intentionally slice-local: resolver code does not reach
across slice boundaries.  When additional slices are ported, extend this file
with new context-builder classes rather than adding cross-slice logic inside
existing ones.
"""

from __future__ import annotations

import json
import string
from copy import copy
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from lib.firewall.resolvers import FirewallResolver
from lib.interface.resolvers import InterfaceResolver
from lib.log_utils import get_logger
from lib.policy_options.resolvers import PolicyOptionsResolver
from lib.registry.graph import DataRegistry
from lib.registry.objects import ConfigRoot
from lib.routing.resolvers import RoutingInstanceResolver
from lib.workspace import WorkspaceManager

logger = get_logger(__name__)


# ---------------------------------------------------------------------------
# Interface-slice context builder
# ---------------------------------------------------------------------------


class InterfaceContextBuilder:
    """Builds reporting payloads for the interface slice.

    This is the MX v1 stand-in for srx-manager's ``GlobalResolver`` on the
    two interface methods we actually need: ``resolve_base_interfaces_context``
    and ``resolve_unit_interfaces_context``. Zone/routing-instance joins are
    dropped because those slices are not yet ingested on MX.

    The separation from the exporter keeps the exporter free of graph
    knowledge and gives future slice ports a clean hook to extend context
    without touching export code.
    """

    def __init__(self, registry: DataRegistry):
        self.r = registry
        self.if_resolver = InterfaceResolver(registry)

    def resolve_base_interfaces_context(
        self, config_root_uid: str
    ) -> List[Dict[str, Any]]:
        """Returns every base interface with its graph-confirmed unit refs.

        Follows ``ConfigRoot -> InterfaceConfigRoot -> BaseInterfaceContainer
        -> BaseInterfaceNode`` and attaches a lightweight ``units`` list.
        """
        interface_root_uid = self.if_resolver.get_interface_config_root(
            config_root_uid
        )
        if not interface_root_uid:
            return []

        base_container_uid = self.if_resolver.get_base_container(interface_root_uid)
        if not base_container_uid:
            return []

        results: List[Dict[str, Any]] = []

        for iface_uid in self.if_resolver.get_all_base_interfaces(base_container_uid):
            context = self.if_resolver.hydrate_base_interface(iface_uid)
            if not context:
                continue

            unit_refs: List[Dict[str, str]] = []
            for unit_uid in self.if_resolver.get_units_for_interface(iface_uid):
                unit_node = self.r.storage.get(unit_uid)
                if unit_node:
                    unit_refs.append({"name": unit_node.name, "uid": unit_uid})

            context["units"] = unit_refs
            results.append(context)

        return results

    def resolve_unit_interfaces_context(
        self, config_root_uid: str
    ) -> List[Dict[str, Any]]:
        """Returns every unit interface with its owning base interface ref.

        Zone + routing-instance joins are intentionally omitted because MX
        v1 does not ingest those slices. ``parent_interface`` is surfaced
        directly from graph-confirmed ``interface_unit`` reverse edges.
        """
        interface_root_uid = self.if_resolver.get_interface_config_root(
            config_root_uid
        )
        if not interface_root_uid:
            return []

        unit_container_uid = self.if_resolver.get_unit_container(interface_root_uid)
        if not unit_container_uid:
            return []

        results: List[Dict[str, Any]] = []

        for unit_uid in self.if_resolver.get_all_unit_uids(unit_container_uid):
            context = self.if_resolver.hydrate_unit(unit_uid)
            if not context:
                continue

            parent_uid = self.if_resolver.get_parent_interface(unit_uid)
            if parent_uid:
                parent_node = self.r.storage.get(parent_uid)
                context["parent_interface"] = {
                    "name": getattr(parent_node, "name", None),
                    "uid": parent_uid,
                }
            else:
                context["parent_interface"] = None

            results.append(context)

        return results


# ---------------------------------------------------------------------------
# Routing-slice context builder
# ---------------------------------------------------------------------------


class RoutingContextBuilder:
    """Builds reporting payloads for the routing slice.

    Tier 1 only — intra-slice, UID-based, no cross-slice joins.
    """

    def __init__(self, registry: DataRegistry):
        self.r = registry
        self.ri_resolver = RoutingInstanceResolver(registry)

    def resolve_routing_instances_context(
        self, config_root_uid: str
    ) -> List[Dict[str, Any]]:
        """Returns a list of routing instance payloads with static routes and BGP."""
        results: List[Dict[str, Any]] = []

        for ri_uid in self.ri_resolver.get_instances(config_root_uid):
            ri = self.ri_resolver.hydrate_routing_instance(ri_uid)
            if not ri:
                continue

            # Interface unit names attached to this RI
            ri["interfaces"] = [
                getattr(self.r.storage.get(u), "name", u)
                for u in self.ri_resolver.get_instance_interfaces(ri_uid)
            ]

            # Static routes
            ri["static_routes"] = [
                self.ri_resolver.hydrate_static_route(r_uid)
                for r_uid in self.ri_resolver.get_ordered_static_routes(ri_uid)
                if self.ri_resolver.hydrate_static_route(r_uid)
            ]

            # BGP groups + neighbors
            bgp_groups: List[Dict[str, Any]] = []
            for g_uid in self.ri_resolver.get_bgp_groups(ri_uid):
                grp = self.ri_resolver.hydrate_bgp_group(g_uid)
                if not grp:
                    continue
                grp["import_policies"] = [
                    getattr(self.r.storage.get(p), "name", p)
                    for p in self.ri_resolver.get_bgp_policies(g_uid, "import_policy")
                ]
                grp["export_policies"] = [
                    getattr(self.r.storage.get(p), "name", p)
                    for p in self.ri_resolver.get_bgp_policies(g_uid, "export_policy")
                ]
                grp["neighbors"] = [
                    self.ri_resolver.hydrate_bgp_neighbor(n_uid)
                    for n_uid in self.ri_resolver.get_bgp_neighbors(g_uid)
                    if self.ri_resolver.hydrate_bgp_neighbor(n_uid)
                ]
                bgp_groups.append(grp)
            ri["bgp_groups"] = bgp_groups

            results.append(ri)

        return results



# ---------------------------------------------------------------------------
# Policy-options context builder
# ---------------------------------------------------------------------------


class PolicyOptionsContextBuilder:
    """Builds reporting payloads for the policy-options slice."""

    def __init__(self, registry: DataRegistry):
        self.po_resolver = PolicyOptionsResolver(registry)

    def resolve_policy_options_context(self, config_root_uid: str) -> Dict[str, Any]:
        """Returns the policy-options payload (prefix-lists, as-paths, statements)."""
        root_uid = self.po_resolver.get_root(config_root_uid)
        if not root_uid:
            return {
                "prefix_lists": [],
                "as_paths": [],
                "policy_statements": [],
            }

        prefix_container_uid = self.po_resolver.get_prefix_list_container(root_uid)
        as_path_container_uid = self.po_resolver.get_as_path_container(root_uid)
        policy_container_uid = self.po_resolver.get_policy_statement_container(root_uid)

        prefix_lists = []
        if prefix_container_uid:
            for uid in self.po_resolver.get_prefix_lists(prefix_container_uid):
                hydrated = self.po_resolver.hydrate_prefix_list(uid)
                if hydrated is not None:
                    prefix_lists.append(hydrated)

        as_paths = []
        if as_path_container_uid:
            for uid in self.po_resolver.get_as_paths(as_path_container_uid):
                hydrated = self.po_resolver.hydrate_as_path(uid)
                if hydrated is not None:
                    as_paths.append(hydrated)

        policy_statements = []
        if policy_container_uid:
            for uid in self.po_resolver.get_policy_statements(policy_container_uid):
                hydrated = self.po_resolver.hydrate_policy_statement(uid, include_terms=True)
                if hydrated is not None:
                    policy_statements.append(hydrated)

        return {
            "prefix_lists": prefix_lists,
            "as_paths": as_paths,
            "policy_statements": policy_statements,
        }


# ---------------------------------------------------------------------------
# Firewall context builder
# ---------------------------------------------------------------------------


class FirewallContextBuilder:
    """Builds reporting payloads for the firewall slice.

    Tier-1 discipline is respected in FirewallResolver: cross-slice UIDs
    (PrefixList, PortList) are returned as raw UID lists.  This builder
    dereferences them to node names by accessing registry storage directly,
    which is the permitted cross-slice read pattern for context builders
    (invariant 10.2).
    """

    def __init__(self, registry: DataRegistry):
        self.r = registry
        self.fw_resolver = FirewallResolver(registry)

    def resolve_firewall_context(
        self, config_root_uid: str
    ) -> List[Dict[str, Any]]:
        """Returns a list of filter payloads with fully resolved term data."""
        root_uid = self.fw_resolver.get_root(config_root_uid)
        if not root_uid:
            return []

        container_uid = self.fw_resolver.get_filter_container(root_uid)
        if not container_uid:
            return []

        results: List[Dict[str, Any]] = []
        for filter_uid in self.fw_resolver.get_filters(container_uid):
            fw_filter = self.fw_resolver.hydrate_filter(filter_uid)
            if not fw_filter:
                continue

            terms: List[Dict[str, Any]] = []
            for term_uid in self.fw_resolver.get_ordered_terms(filter_uid):
                term = self.fw_resolver.hydrate_term(term_uid)
                if not term:
                    continue

                # Dereference cross-slice UIDs to name + prefix lists
                term["source_prefix_lists"] = [
                    {
                        "name": getattr(self.r.storage.get(u), "name", u),
                        "uid": u,
                        "prefixes": list(getattr(self.r.storage.get(u), "prefixes", [])),
                    }
                    for u in term.pop("source_prefix_list_uids", [])
                ]
                term["destination_prefix_lists"] = [
                    {
                        "name": getattr(self.r.storage.get(u), "name", u),
                        "uid": u,
                        "prefixes": list(getattr(self.r.storage.get(u), "prefixes", [])),
                    }
                    for u in term.pop("destination_prefix_list_uids", [])
                ]
                term["source_port_lists"] = [
                    {"name": getattr(self.r.storage.get(u), "name", u), "uid": u}
                    for u in term.pop("source_port_list_uids", [])
                ]
                term["destination_port_lists"] = [
                    {"name": getattr(self.r.storage.get(u), "name", u), "uid": u}
                    for u in term.pop("destination_port_list_uids", [])
                ]
                terms.append(term)

            fw_filter["terms"] = terms
            results.append(fw_filter)

        return results


# ---------------------------------------------------------------------------
# JSON exporter
# ---------------------------------------------------------------------------


class JSONConfigExporter:
    """Writes the interface slice to a deterministic JSON payload."""

    def __init__(self, ws_manager: WorkspaceManager, registry: DataRegistry):
        self.ws_manager = ws_manager
        self.registry = registry
        self.json_payload: Dict[str, Any] = {}
        self.context_builder = InterfaceContextBuilder(registry)
        self.routing_builder = RoutingContextBuilder(registry)
        self.policy_builder = PolicyOptionsContextBuilder(registry)
        self.firewall_builder = FirewallContextBuilder(registry)

    def export_to_json(
        self, config_root_uid: str, output_filename: str | None = None
    ) -> Path:
        """Builds the payload and writes it to disk.

        Args:
            config_root_uid: UID of the ``ConfigRoot`` anchor.
            output_filename: Optional override; defaults to a file named
                after the ConfigRoot UID in the current directory.

        Returns:
            The path the JSON was written to.
        """
        self.build_json_payload(config_root_uid)

        out_path = (
            Path(output_filename)
            if output_filename
            else Path(f"{config_root_uid}.json")
        )
        out_path.write_text(
            json.dumps(self.json_payload, indent=2), encoding="utf-8"
        )
        return out_path

    def build_json_payload(self, config_root_uid: str) -> None:
        """Assembles the payload in-place on ``self.json_payload``."""
        root = self.registry.storage.get(config_root_uid)
        if not isinstance(root, ConfigRoot):
            raise TypeError(
                f"config_root_uid is not a ConfigRoot: {config_root_uid}"
            )

        payload: Dict[str, Any] = {
            "source": {
                "file": Path(root.file_path).name if root.file_path else root.name,
                "sha256": root.metadata.get("sha256"),
                "xml_root": root.metadata.get("root_tag"),
            },
            "interfaces": {
                "base_interfaces": self.context_builder.resolve_base_interfaces_context(
                    config_root_uid
                ),
                "unit_interfaces": self.context_builder.resolve_unit_interfaces_context(
                    config_root_uid
                ),
            },
            "policy_options": self.policy_builder.resolve_policy_options_context(
                config_root_uid
            ),
            "firewall": {
                "filters": self.firewall_builder.resolve_firewall_context(
                    config_root_uid
                ),
            },
            "routing": {
                "instances": self.routing_builder.resolve_routing_instances_context(
                    config_root_uid
                ),
            },
        }

        self.json_payload = payload


# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------


class DocumentTheme:
    """Centralized styling for the Excel export.

    Kept intentionally minimal: a single theme with alternating row fills
    plus a top-header style. More knobs can be added when additional export
    targets need different palettes.
    """

    def __init__(
        self,
        global_font: str = "Tahoma",
        top_header_bg: str = "FFFFFF",
        top_header_font_color: str = "000000",
        top_header_size: int = 12,
        even_stripe_color: str = "F0F8FF",
        odd_stripe_color: str = "D1E8FF",
    ):
        self.global_font = global_font

        self.top_fill = PatternFill(
            start_color=top_header_bg, end_color=top_header_bg, fill_type="solid"
        )
        self.top_font = Font(
            name=global_font,
            color=top_header_font_color,
            size=top_header_size,
            bold=True,
        )

        self.even_fill = PatternFill(
            start_color=even_stripe_color,
            end_color=even_stripe_color,
            fill_type="solid",
        )
        self.odd_fill = PatternFill(
            start_color=odd_stripe_color,
            end_color=odd_stripe_color,
            fill_type="solid",
        )
        self.data_font = Font(name=global_font, size=10)


class ExcelFormatter:
    """Static helpers for consistent openpyxl styling across sheets."""

    ALIGN_CENTER_LEFT_WRAP = Alignment(
        vertical="center", horizontal="left", wrap_text=True
    )
    ALIGN_BOTTOM_LEFT_WRAP = Alignment(
        vertical="bottom", horizontal="left", wrap_text=True
    )
    ALIGN_CENTER_CENTER_NO_WRAP = Alignment(
        vertical="center", horizontal="center", wrap_text=False
    )

    @staticmethod
    def setup_grid(ws, column_config: Dict[str, int], theme: DocumentTheme) -> None:
        """Writes the header row and sets column widths."""
        letters = list(string.ascii_uppercase) + [
            f"{a}{b}"
            for a in string.ascii_uppercase
            for b in string.ascii_uppercase
        ]

        for i, (name, width) in enumerate(column_config.items()):
            col_letter = letters[i]
            cell = ws.cell(row=1, column=i + 1, value=name)
            cell.fill = theme.top_fill
            cell.font = theme.top_font
            cell.alignment = ExcelFormatter.ALIGN_BOTTOM_LEFT_WRAP
            ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = "B2"

    @staticmethod
    def apply_row_theme(ws, row_index: int, num_cols: int, theme: DocumentTheme) -> None:
        """Applies the alternating stripe fill to a data row."""
        fill = theme.even_fill if row_index % 2 == 0 else theme.odd_fill
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_index, column=col)
            cell.fill = fill
            cell.font = theme.data_font
            cell.alignment = ExcelFormatter.ALIGN_CENTER_LEFT_WRAP

    @staticmethod
    def center_align_column(ws, col_index: int, start_row: int = 2) -> None:
        """Horizontally centers every cell in a column from ``start_row``."""
        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_index)
            new_alignment = copy(cell.alignment)
            new_alignment.horizontal = "center"
            cell.alignment = new_alignment

    @staticmethod
    def format_list(values: Any) -> str:
        """Flattens a list-of-strings into a newline-separated cell value.

        Non-list values are stringified as-is. ``None`` becomes the empty
        string so openpyxl does not render literal ``None``.
        """
        if values is None:
            return ""
        if isinstance(values, list):
            return "\n".join(str(v) for v in values)
        return str(values)


# ---------------------------------------------------------------------------
# Excel exporter
# ---------------------------------------------------------------------------


class ExcelConfigExporter:
    """Renders the interface slice into a two-sheet audit workbook."""

    def __init__(self, ws_manager: WorkspaceManager, registry: DataRegistry):
        self.ws_manager = ws_manager
        self.registry = registry

    def export_to_excel(
        self, config_root_uid: str, output_filename: str | None = None
    ) -> Path:
        """Builds the JSON payload in-memory, then writes the flattened sheets."""
        json_exporter = JSONConfigExporter(self.ws_manager, self.registry)
        json_exporter.build_json_payload(config_root_uid)
        data = json_exporter.json_payload

        out_path = (
            Path(output_filename)
            if output_filename
            else Path(f"{config_root_uid}.xlsx")
        )

        theme = DocumentTheme(global_font="Tahoma")

        # Using ExcelWriter in ``w`` mode lets us add our own sheets and
        # skip pandas' default ``Sheet1``. We create the sheets by hand
        # because the rows need per-row theming, not plain ``to_excel``.
        with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
            self._write_base_interfaces_sheet(
                writer, data.get("interfaces", {}).get("base_interfaces", []), theme
            )
            self._write_unit_interfaces_sheet(
                writer, data.get("interfaces", {}).get("unit_interfaces", []), theme
            )
            self._write_prefix_lists_sheet(
                writer, data.get("policy_options", {}).get("prefix_lists", []), theme
            )
            self._write_as_paths_sheet(
                writer, data.get("policy_options", {}).get("as_paths", []), theme
            )
            self._write_policy_statements_sheet(
                writer, data.get("policy_options", {}).get("policy_statements", []), theme
            )
            self._write_firewall_filters_sheet(
                writer, data.get("firewall", {}).get("filters", []), theme
            )
            self._write_routing_instances_sheet(
                writer, data.get("routing", {}).get("instances", []), theme
            )
            self._write_static_routes_sheet(
                writer, data.get("routing", {}).get("instances", []), theme
            )

            # pandas/openpyxl require at least one visible sheet; remove
            # the default empty one if it ended up in the workbook.
            if "Sheet" in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
                del writer.book["Sheet"]

        return out_path

    # ------------------------------------------------------------------
    # Per-sheet writers
    # ------------------------------------------------------------------

    def _write_base_interfaces_sheet(
        self, writer: pd.ExcelWriter, base_data: List[Dict[str, Any]], theme: DocumentTheme
    ) -> None:
        ws = writer.book.create_sheet("BaseInts")
        col_config = {
            "Interface": 22,
            "Kind": 12,
            "Parent/Members": 42,
            "Tagging": 22,
            "Encapsulation": 20,
            "MC-AE": 32,
            "LACP": 18,
            "Description": 40,
            "UID": 40,
        }
        ExcelFormatter.setup_grid(ws, col_config, theme)
        num_cols = len(col_config)

        current_row = 2
        for base in base_data:
            ExcelFormatter.apply_row_theme(ws, current_row, num_cols, theme)

            member_of_parent = base.get("member_of_parent")
            member_interfaces = base.get("member_interfaces") or []

            if member_of_parent:
                parent_info = f"Member of: {member_of_parent}"
            elif member_interfaces:
                names = [
                    m.get("name", "")
                    for m in member_interfaces
                    if isinstance(m, dict)
                ]
                parent_info = "Members: " + ", ".join(n for n in names if n)
            else:
                parent_info = ""

            tagging_bits = []
            if base.get("flexible_vlan_tagging"):
                tagging_bits.append("flexible-vlan-tagging")
            elif base.get("vlan_tagging"):
                tagging_bits.append("vlan-tagging")
            if base.get("native_vlan_id") is not None:
                tagging_bits.append(f"native-vlan={base['native_vlan_id']}")
            tagging_text = ", ".join(tagging_bits)

            mc_ae = base.get("mc_ae") or {}
            mc_ae_bits = []
            if mc_ae:
                if mc_ae.get("mc_ae_id") is not None:
                    mc_ae_bits.append(f"id={mc_ae['mc_ae_id']}")
                if mc_ae.get("redundancy_group") is not None:
                    mc_ae_bits.append(f"rg={mc_ae['redundancy_group']}")
                if mc_ae.get("chassis_id") is not None:
                    mc_ae_bits.append(f"chassis={mc_ae['chassis_id']}")
                if mc_ae.get("mode"):
                    mc_ae_bits.append(f"mode={mc_ae['mode']}")
                if mc_ae.get("status_control"):
                    mc_ae_bits.append(f"status={mc_ae['status_control']}")
            mc_ae_text = ", ".join(mc_ae_bits)

            lacp_bits = []
            if base.get("lacp_mode"):
                lacp_bits.append(base["lacp_mode"])
            if base.get("lacp_periodic"):
                lacp_bits.append(f"periodic={base['lacp_periodic']}")
            if base.get("minimum_links") is not None:
                lacp_bits.append(f"min-links={base['minimum_links']}")
            lacp_text = ", ".join(lacp_bits)

            row_values = [
                base.get("name"),
                base.get("kind"),
                parent_info,
                tagging_text,
                base.get("encapsulation") or "",
                mc_ae_text,
                lacp_text,
                base.get("description") or "",
                base.get("uid"),
            ]

            for c_idx, val in enumerate(row_values, 1):
                ws.cell(row=current_row, column=c_idx, value=val)
            current_row += 1

        ExcelFormatter.center_align_column(ws, col_index=2)

    def _write_unit_interfaces_sheet(
        self, writer: pd.ExcelWriter, unit_data: List[Dict[str, Any]], theme: DocumentTheme
    ) -> None:
        ws = writer.book.create_sheet("UnitInts")
        col_config = {
            "Interface": 22,
            "Parent": 18,
            "Vlan ID": 10,
            "L2?": 6,
            "Encapsulation": 18,
            "Addresses": 32,
            "v4 Address Context": 28,
            "v6 Address Context": 28,
            "Description": 40,
            "UID": 40,
        }
        ExcelFormatter.setup_grid(ws, col_config, theme)
        num_cols = len(col_config)

        current_row = 2
        for unit in unit_data:
            ExcelFormatter.apply_row_theme(ws, current_row, num_cols, theme)

            parent_ref = unit.get("parent_interface") or {}
            parent_name = (
                parent_ref.get("name") if isinstance(parent_ref, dict) else None
            ) or unit.get("parent_name", "")

            row_values = [
                unit.get("name"),
                parent_name,
                unit.get("vlan_id"),
                "yes" if unit.get("is_l2") else "",
                unit.get("encapsulation") or "",
                ExcelFormatter.format_list(unit.get("addresses", [])),
                ExcelFormatter.format_list(unit.get("v4_address_context", [])),
                ExcelFormatter.format_list(unit.get("v6_address_context", [])),
                unit.get("description") or "",
                unit.get("uid"),
            ]

            for c_idx, val in enumerate(row_values, 1):
                ws.cell(row=current_row, column=c_idx, value=val)
            current_row += 1

        ExcelFormatter.center_align_column(ws, col_index=3)
        ExcelFormatter.center_align_column(ws, col_index=4)

    def _write_prefix_lists_sheet(
        self,
        writer: pd.ExcelWriter,
        prefix_lists: List[Dict[str, Any]],
        theme: DocumentTheme,
    ) -> None:
        ws = writer.book.create_sheet("PrefixLists")
        col_config = {
            "Prefix List": 30,
            "Prefixes": 70,
            "UID": 40,
        }
        ExcelFormatter.setup_grid(ws, col_config, theme)
        num_cols = len(col_config)
        current_row = 2
        for prefix_list in prefix_lists:
            ExcelFormatter.apply_row_theme(ws, current_row, num_cols, theme)
            row_values = [
                prefix_list.get("name"),
                ExcelFormatter.format_list(prefix_list.get("prefixes", [])),
                prefix_list.get("uid"),
            ]
            for c_idx, val in enumerate(row_values, 1):
                ws.cell(row=current_row, column=c_idx, value=val)
            current_row += 1

    def _write_as_paths_sheet(
        self,
        writer: pd.ExcelWriter,
        as_paths: List[Dict[str, Any]],
        theme: DocumentTheme,
    ) -> None:
        ws = writer.book.create_sheet("AsPaths")
        col_config = {
            "AS Path": 30,
            "Regex": 70,
            "UID": 40,
        }
        ExcelFormatter.setup_grid(ws, col_config, theme)
        num_cols = len(col_config)
        current_row = 2
        for as_path in as_paths:
            ExcelFormatter.apply_row_theme(ws, current_row, num_cols, theme)
            row_values = [
                as_path.get("name"),
                as_path.get("path"),
                as_path.get("uid"),
            ]
            for c_idx, val in enumerate(row_values, 1):
                ws.cell(row=current_row, column=c_idx, value=val)
            current_row += 1

    def _write_policy_statements_sheet(
        self,
        writer: pd.ExcelWriter,
        policy_statements: List[Dict[str, Any]],
        theme: DocumentTheme,
    ) -> None:
        ws = writer.book.create_sheet("PolicyStatements")
        col_config = {
            "Policy Statement": 30,
            "Term": 22,
            "Protocols": 22,
            "Prefix Lists": 30,
            "AS Paths": 30,
            "Route Filters": 40,
            "Actions": 30,
            "Next Hop": 20,
            "UID": 40,
        }
        ExcelFormatter.setup_grid(ws, col_config, theme)
        num_cols = len(col_config)
        current_row = 2
        for statement in policy_statements:
            terms = statement.get("terms") or [None]
            for term in terms:
                ExcelFormatter.apply_row_theme(ws, current_row, num_cols, theme)
                prefix_list_names = []
                as_path_names = []
                route_filters = []
                if isinstance(term, dict):
                    prefix_list_names = [ref.get("name", "") for ref in term.get("prefix_lists", []) if isinstance(ref, dict)]
                    as_path_names = [ref.get("name", "") for ref in term.get("as_paths", []) if isinstance(ref, dict)]
                    route_filters = [
                        f"{rf.get('address', '')} {rf.get('match_type', '')}".strip()
                        for rf in term.get("route_filters", [])
                        if isinstance(rf, dict)
                    ]
                row_values = [
                    statement.get("name"),
                    term.get("name") if isinstance(term, dict) else "",
                    ExcelFormatter.format_list(term.get("from_protocols", [])) if isinstance(term, dict) else "",
                    ExcelFormatter.format_list(prefix_list_names),
                    ExcelFormatter.format_list(as_path_names),
                    ExcelFormatter.format_list(route_filters),
                    ExcelFormatter.format_list(term.get("actions", [])) if isinstance(term, dict) else "",
                    term.get("next_hop") if isinstance(term, dict) else "",
                    statement.get("uid"),
                ]
                for c_idx, val in enumerate(row_values, 1):
                    ws.cell(row=current_row, column=c_idx, value=val)
                current_row += 1

    def _write_firewall_filters_sheet(
        self,
        writer: pd.ExcelWriter,
        filters: List[Dict[str, Any]],
        theme: DocumentTheme,
    ) -> None:
        """One row per term across all filters.

        Inline match conditions (protocols, ports, addresses) are rendered
        compactly.  Protocol entries show ``name(number)`` or just ``number``
        when there is no known name.  Port entries show ``name(number)``,
        ``number``, or ``low-high`` for ranges.  Cross-slice prefix-list and
        port-list names are newline-joined.
        """
        ws = writer.book.create_sheet("FirewallFilters")
        col_config = {
            "Filter": 30,
            "AF": 8,
            "Term": 28,
            "Protocols": 20,
            "Src Addresses": 34,
            "Dst Addresses": 34,
            "Src Ports": 22,
            "Dst Ports": 22,
            "Ports (either)": 18,
            "ICMP Types": 18,
            "TCP-Est": 8,
            "Src Prefix Lists": 28,
            "Dst Prefix Lists": 28,
            "Actions": 18,
            "Count": 20,
            "UID": 40,
        }
        ExcelFormatter.setup_grid(ws, col_config, theme)
        num_cols = len(col_config)
        current_row = 2

        for fw_filter in filters:
            filter_name = fw_filter.get("name", "")
            af = fw_filter.get("address_family", "")
            for term in fw_filter.get("terms") or []:
                ExcelFormatter.apply_row_theme(ws, current_row, num_cols, theme)

                row_values = [
                    filter_name,
                    af,
                    term.get("name", ""),
                    ExcelFormatter.format_list(
                        [self._fmt_protocol(p) for p in (term.get("protocols") or [])]
                    ),
                    ExcelFormatter.format_list(term.get("source_addresses") or []),
                    ExcelFormatter.format_list(term.get("destination_addresses") or []),
                    ExcelFormatter.format_list(
                        [self._fmt_port(p) for p in (term.get("source_ports") or [])]
                    ),
                    ExcelFormatter.format_list(
                        [self._fmt_port(p) for p in (term.get("destination_ports") or [])]
                    ),
                    ExcelFormatter.format_list(
                        [self._fmt_port(p) for p in (term.get("ports") or [])]
                    ),
                    ExcelFormatter.format_list(
                        [self._fmt_icmp_type(t) for t in (term.get("icmp_types") or [])]
                    ),
                    "yes" if term.get("tcp_established") else "",
                    ExcelFormatter.format_list(
                        [self._fmt_prefix_list_ref(ref) for ref in (term.get("source_prefix_lists") or [])]
                    ),
                    ExcelFormatter.format_list(
                        [self._fmt_prefix_list_ref(ref) for ref in (term.get("destination_prefix_lists") or [])]
                    ),
                    ExcelFormatter.format_list(term.get("actions") or []),
                    term.get("count") or "",
                    term.get("uid", ""),
                ]
                for c_idx, val in enumerate(row_values, 1):
                    ws.cell(row=current_row, column=c_idx, value=val)
                current_row += 1

        # Center-align compact columns
        ExcelFormatter.center_align_column(ws, col_index=2)   # AF
        ExcelFormatter.center_align_column(ws, col_index=11)  # TCP-Est

    @staticmethod
    def _fmt_prefix_list_ref(ref: Dict[str, Any]) -> str:
        """Renders a prefix-list reference with its member CIDRs expanded.

        Example output::

            AWS_US_VA
            [
            1.2.150.0/24
            1.178.4.0/24
            ]
        """
        name = ref.get("name", "")
        prefixes = ref.get("prefixes") or []
        if not prefixes:
            return name
        return f"{name}\n[\n" + "\n".join(prefixes) + "\n]"

    @staticmethod
    def _fmt_protocol(p: Dict[str, Any]) -> str:
        """Renders a normalized protocol dict as a readable string."""
        name = p.get("canonical_name")
        number = p.get("number")
        if name and number is not None:
            return f"{name}({number})"
        if number is not None:
            return str(number)
        return p.get("raw", "")

    @staticmethod
    def _fmt_port(p: Dict[str, Any]) -> str:
        """Renders a normalized port dict as a readable string."""
        kind = p.get("kind", "single")
        if kind == "range":
            return f"{p.get('low')}-{p.get('high')}"
        name = p.get("canonical_name")
        value = p.get("value")
        if name and value is not None:
            return f"{name}({value})"
        if value is not None:
            return str(value)
        return p.get("raw", "")

    @staticmethod
    def _fmt_icmp_type(t: Dict[str, Any]) -> str:
        """Renders a normalized ICMP type dict as a readable string."""
        name = t.get("canonical_name")
        number = t.get("number")
        if name and number is not None:
            return f"{name}({number})"
        if number is not None:
            return str(number)
        return t.get("raw", "")

    def _write_routing_instances_sheet(
        self,
        writer: pd.ExcelWriter,
        instances: List[Dict[str, Any]],
        theme: DocumentTheme,
    ) -> None:
        ws = writer.book.create_sheet("RoutingInstances")
        col_config = {
            "Instance": 30,
            "Type": 18,
            "Interfaces": 40,
            "BGP Groups": 30,
            "Static Routes": 8,
            "Description": 40,
            "UID": 40,
        }
        ExcelFormatter.setup_grid(ws, col_config, theme)
        num_cols = len(col_config)

        current_row = 2
        for ri in instances:
            ExcelFormatter.apply_row_theme(ws, current_row, num_cols, theme)

            iface_names = ri.get("interfaces") or []
            bgp_group_names = [
                g.get("name", "") for g in (ri.get("bgp_groups") or [])
            ]

            row_values = [
                ri.get("name"),
                ri.get("instance_type"),
                ExcelFormatter.format_list(iface_names),
                ExcelFormatter.format_list(bgp_group_names),
                len(ri.get("static_routes") or []),
                ri.get("description") or "",
                ri.get("uid"),
            ]
            for c_idx, val in enumerate(row_values, 1):
                ws.cell(row=current_row, column=c_idx, value=val)
            current_row += 1

        ExcelFormatter.center_align_column(ws, col_index=2)
        ExcelFormatter.center_align_column(ws, col_index=5)

    def _write_static_routes_sheet(
        self,
        writer: pd.ExcelWriter,
        instances: List[Dict[str, Any]],
        theme: DocumentTheme,
    ) -> None:
        """Flattens all static routes from all routing instances into one sheet."""
        ws = writer.book.create_sheet("StaticRoutes")
        col_config = {
            "Routing Instance": 30,
            "Destination": 22,
            "Next Hops": 40,
            "Flags": 14,
            "Preference": 12,
            "Tag": 8,
            "Description": 40,
        }
        ExcelFormatter.setup_grid(ws, col_config, theme)
        num_cols = len(col_config)

        current_row = 2
        for ri in instances:
            ri_name = ri.get("name", "")
            for route in ri.get("static_routes") or []:
                ExcelFormatter.apply_row_theme(ws, current_row, num_cols, theme)

                nh_lines: List[str] = []
                for nh in route.get("next_hops") or []:
                    kind = nh.get("kind", "")
                    if kind == "ip":
                        val = nh.get("ip_address", "")
                    elif kind == "interface":
                        val = nh.get("interface", "")
                    elif kind == "next-table":
                        val = f"next-table {nh.get('next_table', '')}"
                    elif kind in ("discard", "reject"):
                        val = kind
                    else:
                        val = str(nh)
                    suffix = ""
                    if nh.get("qualified"):
                        parts = []
                        if nh.get("preference") is not None:
                            parts.append(f"pref={nh['preference']}")
                        if nh.get("metric") is not None:
                            parts.append(f"metric={nh['metric']}")
                        if parts:
                            suffix = f" ({', '.join(parts)})"
                    nh_lines.append(f"{val}{suffix}")

                flags = []
                if route.get("discard"):
                    flags.append("discard")
                if route.get("reject"):
                    flags.append("reject")

                row_values = [
                    ri_name,
                    route.get("destination"),
                    "\n".join(nh_lines),
                    ", ".join(flags),
                    route.get("preference"),
                    route.get("tag"),
                    route.get("description") or "",
                ]
                for c_idx, val in enumerate(row_values, 1):
                    ws.cell(row=current_row, column=c_idx, value=val)
                current_row += 1

        ExcelFormatter.center_align_column(ws, col_index=5)
        ExcelFormatter.center_align_column(ws, col_index=6)


# end of lib/exporter.py
